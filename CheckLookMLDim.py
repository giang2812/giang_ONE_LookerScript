import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    root_folder = 'Looker'
    directory = 'D:/Aki Work/CyberLogitec/ONE Project/Looker/LOOKML_one_ap_spoke/'
    parameter_hierarchy = ['hidden', 'view_label', 'group_label', 'group_item_label', 'label', 'type','description','sql_distinct_key','sql']
    results = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.view.lkml'):
                full_file_path = os.path.join(root, file)
                relative_folder = root.split(root_folder)[-1]
                if relative_folder == '':
                    folder = ''
                else:
                    folder = "/" + relative_folder.lstrip('/').replace('\\', '/') + "/"
                with open(full_file_path, 'r') as f:
                    content = f.read()

                view_blocks = re.split(r'(?=view:\s*[\w+]+\s*{)', content)
                for view_block in view_blocks:
                    if "view:" in view_block:
                        view_name = re.search(r'view:\s*([\w+]+)', view_block)
                        if view_name:
                            view_name = view_name.group(1)
                            dimensions = re.findall(r'(dimension:\s*\w+\s*{[^}]*})', view_block)
                            dimension_groups = re.findall(r'(dimension_group:\s*\w+\s*{[^}]*})', view_block)
                            measures = re.findall(r'(measure:\s*\w+\s*{[^}]*})', view_block)

                            # Check dimensions
                            for dim in dimensions:
                                dim_name = re.search(r'dimension:\s*(\w+)', dim).group(1)
                                parameters = re.findall(r'\s*(\w+):\s*', dim)
                                check_parameter_order(file, folder, view_name, dim_name, parameters, parameter_hierarchy,
                                                      "dimension", results)

                            # Check dimension groups
                            for dim_group in dimension_groups:
                                dim_group_name = re.search(r'dimension_group:\s*(\w+)', dim_group).group(1)
                                parameters = re.findall(r'\s*(\w+):\s*', dim_group)
                                check_parameter_order(file, folder, view_name, dim_group_name, parameters, parameter_hierarchy,
                                                      "dimension_group", results)

                            # Check measures
                            for measure in measures:
                                measure_name = re.search(r'measure:\s*(\w+)', measure).group(1)
                                parameters = re.findall(r'\s*(\w+):\s*', measure)
                                check_parameter_order(file, folder, view_name, measure_name, parameters, parameter_hierarchy,
                                                      "measure", results)

    # Create a pandas DataFrame from the results
    df = pd.DataFrame(results, columns=["Folder", "Filename", "View", "Field Name", "Field Type", "Expected Order", "Current Order"])

    # Save the DataFrame to an Excel file
    export_to_excel(df, "parameter_order_results.xlsx")


def check_parameter_order(file, folder, view_name, obj_name, parameters, parameter_hierarchy, obj_type, results):
    # Check the order
    correct_order = True
    last_index = -1
    filtered_params = [param for param in parameters if param in parameter_hierarchy]
    for param in filtered_params:
        if param in parameter_hierarchy:
            param_index = parameter_hierarchy.index(param)
            if param_index < last_index:
                correct_order = False
                break
            last_index = param_index

    if not correct_order:
        parameter_order_str = ', '.join([f"{param}" for param in parameter_hierarchy if param in filtered_params])
        current_order_str = ', '.join([f"{param}" for param in filtered_params])
        results.append((folder, file, view_name, obj_name, obj_type, parameter_order_str, current_order_str))




def export_to_excel(df, filename):
    # Create a workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Write the column headers
    for col_num, header in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    # Write the data rows
    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}{row_num}"] = value

    # Auto-fit column widths
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Center-align the content in all cells
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    workbook.save(filename)


if __name__ == "__main__":
    main()