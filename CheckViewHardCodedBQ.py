import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    root_folder = 'Looker'
    directory = 'D:/Aki Work/CyberLogitec/ONE Project/Looker/LOOKML_one_ap_spoke/'
    results = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.view.lkml'):
                full_file_path = os.path.join(root, file)
                relative_folder = root.split(root_folder)[-1]
                if relative_folder == '':
                    folder = ''
                else:
                    folder = "/" + relative_folder.lstrip('/').replace('\\', '/') + "/"
                with open(full_file_path, 'r') as f:
                    content = f.read()

                view_blocks = re.split(r'(?=view:\s*[\w+]+\s*{)', content)
                for view_block in view_blocks:
                    if "view:" in view_block:
                        view_name = re.search(r'view:\s*([\w+]+)', view_block)
                        if view_name:
                            view_name = view_name.group(1)
                            sql_table_name = re.search(r'sql_table_name:\s*`([^`]+)`', view_block)
                            if sql_table_name:
                                sql_table_name = sql_table_name.group(1)
                                if "@" not in sql_table_name:
                                    results.append((folder, file, view_name, sql_table_name))

    df = pd.DataFrame(results, columns=["Folder", "Filename", "View", "Hardcoded BQ View"])
    export_to_excel(df, "hardcoded_bq_view_results.xlsx")


def export_to_excel(df, filename):
    workbook = Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}{row_num}"] = value

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(filename)


if __name__ == "__main__":
    main()
