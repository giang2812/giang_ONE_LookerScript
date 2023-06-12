import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    root_folder = 'Looker'
    directory = 'D:/Aki Work/CyberLogitec/ONE Project/Looker/'
    results = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.view.lkml'):
                full_file_path = os.path.join(root, file)
                relative_folder = root.split(root_folder)[-1]
                folder = "/" + relative_folder.lstrip('/').replace('\\', '/') + "/" if relative_folder else ''
                with open(full_file_path, 'r') as f:
                    content = f.read()

                view_blocks = re.split(r'(?=view:\s*[\w+]+\s*{)', content)
                for view_block in view_blocks:
                    if "view:" in view_block:
                        view_name = re.search(r'view:\s*([\w+]+)', view_block).group(1)
                        dimensions = re.findall(r'(dimension:\s*\w+\s*{[^}]*})', view_block)
                        dimension_groups = re.findall(r'(dimension_group:\s*\w+\s*{[^}]*})', view_block)
                        measures = re.findall(r'(measure:\s*\w+\s*{[^}]*})', view_block)

                        for dim in dimensions:
                            dim_name = re.search(r'dimension:\s*(\w+)', dim).group(1)
                            record_all_params(file, folder, view_name, dim_name, "dimension", dim, results)
                        
                        for dim_group in dimension_groups:
                            dim_group_name = re.search(r'dimension_group:\s*(\w+)', dim_group).group(1)
                            record_all_params(file, folder, view_name, dim_group_name, "dimension_group", dim_group, results)
                        
                        for measure in measures:
                            measure_name = re.search(r'measure:\s*(\w+)', measure).group(1)
                            record_all_params(file, folder, view_name, measure_name, "measure", measure, results)
    
    df = pd.DataFrame(results, columns=["Folder", "Filename", "View", "Field Name", "Field Type", "Current Parameter"])
    export_to_excel(df, "parameter_results.xlsx")


def record_all_params(file, folder, view_name, obj_name, obj_type, obj_block, results):
    params = re.findall(r'\s*(\w+):\s*', obj_block)
    params = [p for p in params if p not in ['dimension', 'dimension_group', 'measure']]
    if 'type' in params or 'sql' in params:
        current_params_str = ', '.join(params)
        results.append((folder, file, view_name, obj_name, obj_type, current_params_str))


def export_to_excel(df, filename):
    workbook = Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}{row_num}"] = value

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(filename)


if __name__ == "__main__":
    main()
