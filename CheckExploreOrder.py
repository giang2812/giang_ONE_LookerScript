import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    root_folder = 'Looker'
    directory = 'D:/Aki Work/CyberLogitec/ONE Project/Looker/LOOKML_one_ap_spoke/'
    parameter_hierarchy = ['view_name', 'group_label', 'label', 'description']
    results = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.explore.lkml'):
                full_file_path = os.path.join(root, file)
                relative_folder = root.split(root_folder)[-1]
                if relative_folder == '':
                    folder = ''
                else:
                    folder = "/" + relative_folder.lstrip('/').replace('\\', '/') + "/"
                with open(full_file_path, 'r') as f:
                    content = f.read()

                explore_blocks = re.split(r'(?=explore:\s*[\w+]+\s*{)', content)
                for explore_block in explore_blocks:
                    if "explore:" in explore_block:
                        explore_name = re.search(r'explore:\s*([\w+]+)', explore_block)
                        if explore_name:
                            explore_name = explore_name.group(1)
                            parameters = re.findall(r'\s*(\w+):\s*', explore_block)
                            check_parameter_order(file, folder, explore_name, parameters, parameter_hierarchy, results)

    # Create a pandas DataFrame from the results
    df = pd.DataFrame(results, columns=["Folder", "Filename", "Explore", "Expected Order", "Current Order"])

    # Save the DataFrame to an Excel file
    export_to_excel(df, "explore_parameter_order_results.xlsx")


def check_parameter_order(file, folder, explore_name, parameters, parameter_hierarchy, results):
    # Check the order
    correct_order = True
    last_index = -1
    filtered_params = [param for param in parameters if param in parameter_hierarchy]
    for param in filtered_params:
        if param in parameter_hierarchy:
            param_index = parameter_hierarchy.index(param)
            if param_index < last_index:
                correct_order = False
                break
            last_index = param_index

    if not correct_order:
        parameter_order_str = ', '.join([f"{param}" for param in parameter_hierarchy if param in filtered_params])
        current_order_str = ', '.join([f"{param}" for param in filtered_params])
        results.append((folder, file, explore_name, parameter_order_str, current_order_str))



def export_to_excel(df, filename):
    # Create a workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Write the column headers
    for col_num, header in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    # Write the data rows
    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}{row_num}"] = value

    # Auto-fit column widths
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Center-align the content in all cells
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    workbook.save(filename)


if __name__ == "__main__":
    main()
